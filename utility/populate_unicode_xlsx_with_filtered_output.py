#!/usr/bin/env python3

"""
Purpose:
	Adds or updates existing XML spreadsheet with:
		- Column with "[py: range]" in the heading: Max 256-character named Unicode range blocks,
		- Column with "[py: orig]" in the heading: Original printable non-space characters; single-spaced.
			in the format "U+[lowest code pointin 'py: orig column']-U+[highest code point in 'py: orig column']"
		- Column with "[py: filter_junk]" in the heading: Output from 'filter_1_junk.py', using contents of '[py: orig]' cell as input.
		- Column with "[py: filter_messy]" in the heading: Output from 'filter_2_messy.py', using contents of '[py: filter_junk]' cell as input.
		- Column with "[py: filter_visual]" in the heading: Output from 'filter_3_visual.py', with contents of '[py: filter_messy]' cell as input.

Copyright © 2026 Jim Collier (ID: 1cv◂‡Vᛦ)
Licensed under the GNU General Public License v2.0 or later. Full text at:
	https://spdx.org/licenses/GPL-2.0-or-later.html
SPDX-License-Identifier: GPL-2.0-or-later
"""

import sys
import os
import re
import unicodedata
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
sys.path.insert(0, os.path.join(SCRIPT_DIR, 'include'))

import openpyxl
from generate_unicode_all_grouped_by_block import (
	BLOCKS_DATA, parse_blocks, is_printable, split_chunks, fmt_cp, fmt_range, MAX_ROW
)
from filter_1_junk import extract as filter_junk_extract
from filter_2_messy import extract as filter_messy_extract
from filter_3_visual import extract as filter_visual_extract

XLSX_PATH = os.path.join(SCRIPT_DIR, '..', 'reference', 'unicode_good_base_symbols.xlsx')
MAX_CODEPOINT = 0x20900
REQUIRED_TAGS = ['range', 'orig', 'filter_junk', 'filter_messy', 'filter_visual', 'updated']


def find_py_columns(ws):
	"""Scan header row for [py: XXX] tags, return dict of tag -> column index."""
	cols = {}
	for col_idx in range(1, ws.max_column + 1):
		val = ws.cell(row=1, column=col_idx).value
		if val and '[py:' in str(val):
			m = re.search(r'\[py:\s*(\w+)\]', str(val))
			if m:
				cols[m.group(1)] = col_idx
	return cols


def parse_range_str(s):
	"""Parse 'U+XXXX-U+YYYY' into (start_int, end_int) or None."""
	m = re.match(r'U\+([0-9A-Fa-f]+)\s*-\s*U\+([0-9A-Fa-f]+)', str(s).strip())
	if m:
		return (int(m.group(1), 16), int(m.group(2), 16))
	return None


def build_row_index(ws, range_col):
	"""Build dict: (start, end) -> row_number for all existing ranges.
	   Stops after 100 consecutive empty rows to avoid scanning padding."""
	idx = {}
	empty_run = 0
	for row_idx in range(2, ws.max_row + 1):
		val = ws.cell(row=row_idx, column=range_col).value
		if val:
			empty_run = 0
			parsed = parse_range_str(val)
			if parsed:
				idx[parsed] = row_idx
		else:
			empty_run += 1
			if empty_run > 100:
				break
	return idx


def find_matching_row(chunk_start, chunk_end, block_start, block_end, row_index):
	"""Find existing row matching this chunk. Returns row_number or None."""
	# Exact match on chunk range
	if (chunk_start, chunk_end) in row_index:
		return row_index[(chunk_start, chunk_end)]
	# Subset match: any existing range within block that overlaps this chunk
	for (rs, re_), row in row_index.items():
		if block_start <= rs and re_ <= block_end:
			if rs <= chunk_end and re_ >= chunk_start:
				return row
	return None


def generate_block_chunks():
	"""Generate all (range_str, chunk_start, chunk_end, orig_str, block_start, block_end)
	   tuples for blocks below MAX_CODEPOINT."""
	blocks = parse_blocks(BLOCKS_DATA)
	chunks_out = []
	for start, end, name in blocks:
		if start >= MAX_CODEPOINT:
			continue
		chars = []
		for cp in range(start, end + 1):
			if not is_printable(cp):
				continue
			c = chr(cp)
			if unicodedata.category(c).startswith('Z'):
				continue
			chars.append(c)
		if not chars:
			continue
		for chunk in split_chunks(chars, MAX_ROW):
			chunk_start = ord(chunk[0])
			chunk_end = ord(chunk[-1])
			range_str = fmt_range(chunk_start, chunk_end)
			orig_str = ' '.join(chunk)
			chunks_out.append((range_str, chunk_start, chunk_end, orig_str, start, end))
	return chunks_out


def main():
	print(f"Opening {XLSX_PATH}", file=sys.stderr)
	wb = openpyxl.load_workbook(XLSX_PATH)
	ws = wb.active

	# Discover columns
	py_cols = find_py_columns(ws)
	missing = [t for t in REQUIRED_TAGS if t not in py_cols]
	if missing:
		print(f"Error: missing required column tags: {missing}", file=sys.stderr)
		print(f"Found tags: {py_cols}", file=sys.stderr)
		sys.exit(1)

	col_range   = py_cols['range']
	col_orig    = py_cols['orig']
	col_junk    = py_cols['filter_junk']
	col_messy   = py_cols['filter_messy']
	col_visual  = py_cols['filter_visual']
	col_updated = py_cols['updated']
	print(f"Columns: range={col_range}, orig={col_orig}, filter_junk={col_junk}, filter_messy={col_messy}, filter_visual={col_visual}, updated={col_updated}", file=sys.stderr)

	# Build row index
	row_index = build_row_index(ws, col_range)
	print(f"Found {len(row_index)} existing range rows", file=sys.stderr)

	# Find actual last data row for appending
	next_row = max(row_index.values()) + 1 if row_index else 2

	# Generate block chunks
	chunks = generate_block_chunks()
	print(f"Generated {len(chunks)} block chunks below U+{MAX_CODEPOINT:05X}", file=sys.stderr)

	for i, (range_str, chunk_start, chunk_end, orig_str, block_start, block_end) in enumerate(chunks):
		# Find existing row or allocate new one
		row = find_matching_row(chunk_start, chunk_end, block_start, block_end, row_index)
		is_new = row is None
		if is_new:
			row = next_row
			next_row += 1

		# [py: range] — populate only if empty
		existing_range = ws.cell(row=row, column=col_range).value
		if not existing_range:
			ws.cell(row=row, column=col_range).value = range_str

		# [py: orig] — populate only if empty
		existing_orig = ws.cell(row=row, column=col_orig).value
		if not existing_orig:
			ws.cell(row=row, column=col_orig).value = orig_str
			# Update range based on actual filtered chars
			actual_chars = orig_str.split()
			if actual_chars:
				actual_start = ord(actual_chars[0])
				actual_end = ord(actual_chars[-1])
				ws.cell(row=row, column=col_range).value = fmt_range(actual_start, actual_end)

		# Read current orig for filter inputs
		current_orig = ws.cell(row=row, column=col_orig).value or ''

		# [py: filter_junk] — always recompute
		if current_orig:
			junk_result = filter_junk_extract(current_orig)
			ws.cell(row=row, column=col_junk).value = junk_result
		else:
			junk_result = ''
			ws.cell(row=row, column=col_junk).value = ''

		# [py: filter_messy] — always recompute
		if junk_result:
			messy_result = filter_messy_extract(junk_result)
			ws.cell(row=row, column=col_messy).value = messy_result
		else:
			messy_result = ''
			ws.cell(row=row, column=col_messy).value = ''

		# [py: filter_visual] — always recompute
		if messy_result:
			visual_result = filter_visual_extract(messy_result)
			ws.cell(row=row, column=col_visual).value = visual_result
		else:
			ws.cell(row=row, column=col_visual).value = ''

		# [py: updated] — timestamp this row
		ws.cell(row=row, column=col_updated).value = datetime.now()
		ws.cell(row=row, column=col_updated).number_format = 'YYYY-MM-DD HH:MM:SS'

		status = "NEW" if is_new else "updated"
		print(f"  [{i+1}/{len(chunks)}] Row {row} ({status}): {range_str}", file=sys.stderr)

	# Save
	wb.save(XLSX_PATH)
	print(f"\nSaved {XLSX_PATH}", file=sys.stderr)


if __name__ == '__main__':
	main()

"""
Claude Opus 4.6 instructions:
	Creation 20260507:
		Complete this python script to:
		1. Open a specified Excel spreadsheet.
		2. For each max 256-character named Unicode range block lower than U+20900:
		2.1 Find its existing row in the spreadsheet. It's range may exist as a subset (e.g. U+0021-U+007E). If so, use the subset.
		2.1.1: Search in the column that has the substring "[py: range]" in it.
		2.1.2 If you didn't find a row, add a new one at the bottom. (It will be manually sorted later.)
		2.2 Update/overwrite the new or existing row as follows:
		2.2.1 Add the Range for the block, if it doesn't exist, in the cell that has the substring "[py: range]" in the column header.
		2.2.2 Add the characters for the block, if they don't exist, in the cell that has the substring "[py: orig]" in the column header.
		2.2.2.1 Filter the characters so to remove:
		2.2.2.1.1 Non-printable characters.
		2.2.2.1.2 Space-like characters.
		2.2.2.2 Delimit the characters in the list with a single ASCII space.
		2.2.2.2 After filtering, update the range in the "[py: range]" cell, if necessary (or don't even populate anything until those two values are figured out).
		2.2.3 Populate the cell with "[py: filter_junk]" substring in the heading, with the output from running 'filter_junk.py', using the contents of '[py: orig]' as input.
		2.2.4 Populate the cell with "[py: filter_visual]" substring in the heading, with the output from running 'filter_visual.py', using the contents of '[py: filter_junk]' as input.
		3. Don't overwrite any other cells.
	Update 20260508-090500
		Update Script:
		1. Column with "[py: filter_junk]" in the heading: Output from 'filter_1_junk.py', using contents of '[py: orig]' cell as input.
		2. Column with "[py: filter_messy]" in the heading: Output from 'filter_2_messy.py', using contents of '[py: filter_junk]' cell as input.
		3 Column with "[py: filter_visual]" in the heading: Output from 'filter_3_visual.py', with contents of '[py: filter_messy]' cell as input.
		- You already accomplished the code for Task 1, but it needs its destination column updated.
		- You already accomplished the code for Task 3, but it needs its source column updated.
		- And Task 2, running 'filter_2_messy.py', is new.
	Update 20260508-091300
		Update Script:
		- There is a column with "[py: updated]" in the heading. After updating a row, add the current date/time to that field. The column is defined as "date/time" for display formatting, with the format string "YYYY-mm-DD HH:MM:SS", if that matters.
"""
