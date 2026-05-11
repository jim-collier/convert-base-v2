#!/usr/bin/env python3

"""
Purpose:
	Adds or updates existing XML spreadsheet with:
		- Column with "[py: range]" in the heading: Max 256-character named Unicode range blocks,
		- Column with "[py: orig]" in the heading: Original printable non-space characters; single-spaced.
			in the format "U+[lowest code pointin 'py: orig column']-U+[highest code point in 'py: orig column']"
		- Column with "[py: filter_junk]" in the heading: Output from 'filter_1_junk.py', using contents of '[py: orig]' cell as input.
		- Column with "[py: filter_messy]" in the heading: Output from 'filter_2_messy.py', using contents of '[py: filter_junk]' cell as input.
		- Column with "[py: filter_visual]" in the heading: Output from 'filter_3_visual.py', with contents of '[py: filter_messy]' cell as input.

Copyright © 2026 Jim Collier (ID: 1cv◂‡Vᛦ)
Licensed under the GNU General Public License v2.0 or later. Full text at:
	https://spdx.org/licenses/GPL-2.0-or-later.html
SPDX-License-Identifier: GPL-2.0-or-later
"""

import sys
import os
import re
import gzip
import unicodedata
import xml.etree.ElementTree as ET
from abc import ABC, abstractmethod
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
sys.path.insert(0, os.path.join(SCRIPT_DIR, 'include'))

from generate_unicode_all_grouped_by_block import (
	BLOCKS_DATA, parse_blocks, is_printable, split_chunks, fmt_cp, fmt_range, MAX_ROW
)
from filter_1_junk import extract as filter_junk_extract
from filter_2_messy import extract as filter_messy_extract
from filter_3_visual import extract as filter_visual_extract


SPREADSHEET_NAMES = [
	'unicode_good_base_symbols.ods',
	'unicode_good_base_symbols.gnumeric',
	'unicode_good_base_symbols.xlsx',
]
SPREADSHEET_DIR = os.path.join(SCRIPT_DIR, '..', 'reference')
MAX_CODEPOINT = 0x20900
REQUIRED_TAGS = ['range', 'orig', 'filter_junk', 'filter_messy', 'filter_visual', 'filtered_out', 'updated']

GNM_NS = 'http://www.gnumeric.org/v10.dtd'


## Spreadsheet adapter interface

class SpreadsheetAdapter(ABC):
	"""Abstract interface for reading/writing spreadsheet cells (1-based row/col)."""

	@abstractmethod
	def cell_value(self, row, col):
		"""Return cell value as str, float, or None."""

	@abstractmethod
	def set_cell_value(self, row, col, value):
		"""Set cell value."""

	@abstractmethod
	def set_cell_number_format(self, row, col, fmt):
		"""Set cell display format (e.g. for dates)."""

	@property
	@abstractmethod
	def max_column(self):
		"""Return max column index (1-based)."""

	@property
	@abstractmethod
	def max_row(self):
		"""Return max row index (1-based)."""

	@abstractmethod
	def save(self):
		"""Write changes back to file."""


## XlsxAdapter (openpyxl)

class XlsxAdapter(SpreadsheetAdapter):
	def __init__(self, path):
		try:
			import openpyxl
		except ImportError:
			raise SystemExit("Error: 'openpyxl' is required for .xlsx files. Install it with: pip install openpyxl")
		self._path = path
		self._wb = openpyxl.load_workbook(path)
		self._ws = self._wb.active

	def cell_value(self, row, col):
		return self._ws.cell(row=row, column=col).value

	def set_cell_value(self, row, col, value):
		self._ws.cell(row=row, column=col).value = value

	def set_cell_number_format(self, row, col, fmt):
		self._ws.cell(row=row, column=col).number_format = fmt

	@property
	def max_column(self):
		return self._ws.max_column

	@property
	def max_row(self):
		return self._ws.max_row

	def save(self):
		self._wb.save(self._path)


## GnumericAdapter (gzip + ElementTree)

def _datetime_to_serial(dt):
	"""Convert datetime to Excel serial number (days since 1899-12-30)."""
	epoch = datetime(1899, 12, 30)
	return (dt - epoch).total_seconds() / 86400.0


class GnumericAdapter(SpreadsheetAdapter):
	def __init__(self, path):
		self._path = path
		with gzip.open(path, 'rb') as f:
			self._tree = ET.parse(f)
		self._root = self._tree.getroot()
		# Register namespace to preserve prefix on save
		ET.register_namespace('gnm', GNM_NS)
		# Find first sheet's Cells container
		sheet = self._root.find(f'.//{{{GNM_NS}}}Sheet')
		self._cells_container = sheet.find(f'{{{GNM_NS}}}Cells')
		# Build cell index: (0-based row, 0-based col) -> Element
		self._cells = {}
		self._max_row = 0
		self._max_col = 0
		for cell_el in self._cells_container.findall(f'{{{GNM_NS}}}Cell'):
			r = int(cell_el.get('Row'))
			c = int(cell_el.get('Col'))
			self._cells[(r, c)] = cell_el
			if r > self._max_row:
				self._max_row = r
			if c > self._max_col:
				self._max_col = c

	def cell_value(self, row, col):
		# Convert 1-based to 0-based
		el = self._cells.get((row - 1, col - 1))
		if el is None:
			return None
		vt = el.get('ValueType')
		text = el.text
		if text is None:
			return None
		if vt == '40':  # float
			try:
				return float(text)
			except ValueError:
				return text
		return text

	def set_cell_value(self, row, col, value):
		r, c = row - 1, col - 1
		el = self._cells.get((r, c))
		if el is None:
			el = ET.SubElement(self._cells_container, f'{{{GNM_NS}}}Cell')
			el.set('Row', str(r))
			el.set('Col', str(c))
			self._cells[(r, c)] = el
			if r > self._max_row:
				self._max_row = r
			if c > self._max_col:
				self._max_col = c
		if isinstance(value, datetime):
			el.set('ValueType', '40')
			el.text = str(_datetime_to_serial(value))
		elif isinstance(value, (int, float)):
			el.set('ValueType', '40')
			el.text = str(value)
		else:
			el.set('ValueType', '60')
			el.text = str(value) if value is not None else ''

	def set_cell_number_format(self, row, col, fmt):
		pass  # Format is defined in gnumeric StyleRegion, not per-cell

	@property
	def max_column(self):
		return self._max_col + 1  # convert 0-based to 1-based

	@property
	def max_row(self):
		return self._max_row + 1  # convert 0-based to 1-based

	def save(self):
		xml_bytes = ET.tostring(self._root, xml_declaration=True, encoding='UTF-8')
		with gzip.open(self._path, 'wb') as f:
			f.write(xml_bytes)


## OdsAdapter (odfpy)

class OdsAdapter(SpreadsheetAdapter):
	def __init__(self, path):
		try:
			from odf.opendocument import load
			from odf.table import Table, TableRow, TableCell
			from odf.text import P
		except ImportError:
			raise SystemExit("Error: 'odfpy' is required for .ods files. Install it with: pip install odfpy")
		self._path = path
		self._doc = load(path)
		self._Table = Table
		self._TableRow = TableRow
		self._TableCell = TableCell
		self._P = P
		sheet = self._doc.spreadsheet.getElementsByType(Table)[0]
		self._sheet = sheet
		# Expand rows into indexed grid: (1-based row, 1-based col) -> cell element
		self._cells = {}
		self._row_elements = {}  # 1-based row -> TableRow element
		self._max_row = 0
		self._max_col = 0
		raw_rows = sheet.getElementsByType(TableRow)
		logical_row = 1
		for raw_row in raw_rows:
			row_rep = raw_row.getAttribute('numberrowsrepeated')
			row_rep_n = int(row_rep) if row_rep else 1
			# Only expand up to a reasonable limit for repeated empty rows
			if row_rep_n > 1000:
				break
			for _ in range(row_rep_n):
				self._row_elements[logical_row] = raw_row
				raw_cells = raw_row.getElementsByType(TableCell)
				logical_col = 1
				for raw_cell in raw_cells:
					col_rep = raw_cell.getAttribute('numbercolumnsrepeated')
					col_rep_n = int(col_rep) if col_rep else 1
					# Only store non-empty cells (or first of repeated empties)
					ps = raw_cell.getElementsByType(P)
					has_value = bool(ps) or raw_cell.getAttribute('valuetype')
					if has_value:
						for k in range(col_rep_n):
							self._cells[(logical_row, logical_col + k)] = raw_cell if k == 0 else None
						if logical_col + col_rep_n - 1 > self._max_col:
							self._max_col = logical_col + col_rep_n - 1
					logical_col += col_rep_n
				if logical_row > self._max_row:
					self._max_row = logical_row
				logical_row += 1

	def _get_cell_text(self, cell_el):
		if cell_el is None:
			return None
		ps = cell_el.getElementsByType(self._P)
		if not ps:
			return None
		# Concatenate text from all P elements
		parts = []
		for p in ps:
			# Recursively get all text content
			text = ''
			for node in p.childNodes:
				if hasattr(node, 'data'):
					text += node.data
				elif hasattr(node, '__str__'):
					text += str(node)
			parts.append(text)
		return '\n'.join(parts) if parts else None

	def cell_value(self, row, col):
		cell_el = self._cells.get((row, col))
		if cell_el is None:
			return None
		vtype = cell_el.getAttribute('valuetype')
		if vtype == 'float':
			val = cell_el.getAttribute('value')
			if val is not None:
				try:
					return float(val)
				except ValueError:
					pass
		if vtype == 'date':
			dval = cell_el.getAttribute('datevalue')
			if dval is not None:
				return dval
		return self._get_cell_text(cell_el)

	def _find_cell_in_row(self, row_el, col):
		"""Walk row's child cells to find the element covering logical column `col` (1-based).
		Returns (cell_el, start_col, repeat_count) or (None, None, None) if col is beyond the row."""
		logical_col = 1
		for child in list(row_el.childNodes):
			qname = getattr(child, 'qname', None)
			if not qname or qname[1] != 'table-cell':
				continue
			rep = child.getAttribute('numbercolumnsrepeated')
			rep_n = int(rep) if rep else 1
			if logical_col <= col < logical_col + rep_n:
				return (child, logical_col, rep_n)
			logical_col += rep_n
		return (None, None, None)

	def _clone_cell_attrs(self, source_cell, repeat_count=None):
		"""Create an empty cell with the same style as source_cell."""
		clone = self._TableCell()
		style = source_cell.getAttribute('stylename')
		if style:
			clone.setAttribute('stylename', style)
		if repeat_count is not None and repeat_count > 1:
			clone.setAttribute('numbercolumnsrepeated', str(repeat_count))
		return clone

	def _insert_before(self, parent, new_child, ref_child):
		"""insertBefore that also registers the element with odfpy's document cache."""
		parent.insertBefore(new_child, ref_child)
		# odfpy's insertBefore doesn't register with the document cache,
		# so a later removeChild would fail. Register manually.
		if hasattr(parent, 'ownerDocument') and parent.ownerDocument:
			parent.ownerDocument.rebuild_caches(new_child)

	def set_cell_value(self, row, col, value):
		from odf.table import TableCell, TableRow
		from odf.text import P
		row_el = self._row_elements.get(row)
		if row_el is None:
			# Create a new row
			row_el = TableRow()
			self._sheet.addElement(row_el)
			self._row_elements[row] = row_el

		# Find the actual cell element at this column (even if it's a repeated empty cell)
		found_cell, found_start, found_rep = self._find_cell_in_row(row_el, col)

		# Create new cell, preserving style from the cell we're replacing
		new_cell = TableCell()
		if found_cell is not None:
			old_style = found_cell.getAttribute('stylename')
			if old_style:
				new_cell.setAttribute('stylename', old_style)

		if isinstance(value, datetime):
			iso = value.strftime('%Y-%m-%dT%H:%M:%S')
			new_cell.setAttribute('valuetype', 'date')
			new_cell.setAttribute('datevalue', iso)
			display = value.strftime('%Y-%m-%d %H:%M:%S')
			p = P()
			p.addText(display)
			new_cell.addElement(p)
		elif isinstance(value, (int, float)):
			new_cell.setAttribute('valuetype', 'float')
			new_cell.setAttribute('value', str(value))
			p = P()
			p.addText(str(value))
			new_cell.addElement(p)
		else:
			sval = str(value) if value is not None else ''
			new_cell.setAttribute('valuetype', 'string')
			p = P()
			p.addText(sval)
			new_cell.addElement(p)

		if found_cell is not None:
			if found_rep == 1:
				# Simple 1:1 replacement — no repeat to split
				self._insert_before(row_el, new_cell, found_cell)
				row_el.removeChild(found_cell)
			else:
				# Split a repeated cell group: [before] [new_cell] [after]
				offset = col - found_start
				remaining_after = found_rep - offset - 1
				if offset > 0:
					before = self._clone_cell_attrs(found_cell, offset)
					self._insert_before(row_el, before, found_cell)
				self._insert_before(row_el, new_cell, found_cell)
				if remaining_after > 0:
					after = self._clone_cell_attrs(found_cell, remaining_after)
					self._insert_before(row_el, after, found_cell)
				row_el.removeChild(found_cell)
		else:
			# Column is beyond the current row extent — just append
			row_el.addElement(new_cell)
		self._cells[(row, col)] = new_cell
		if row > self._max_row:
			self._max_row = row
		if col > self._max_col:
			self._max_col = col

	def set_cell_number_format(self, row, col, fmt):
		pass  # Format is defined in ODS styles, not per-cell

	@property
	def max_column(self):
		return self._max_col

	@property
	def max_row(self):
		return self._max_row

	def save(self):
		self._doc.save(self._path)


## Factory

def open_spreadsheet(path):
	"""Open a spreadsheet file and return the appropriate adapter."""
	ext = os.path.splitext(path)[1].lower()
	if ext == '.xlsx':
		return XlsxAdapter(path)
	elif ext == '.gnumeric':
		return GnumericAdapter(path)
	elif ext == '.ods':
		return OdsAdapter(path)
	else:
		raise ValueError(f"Unsupported spreadsheet format: {ext}")


def find_py_columns(adapter):
	"""Scan header row for [py: XXX] tags, return dict of tag -> column index."""
	cols = {}
	for col_idx in range(1, adapter.max_column + 1):
		val = adapter.cell_value(1, col_idx)
		if val and '[py:' in str(val):
			m = re.search(r'\[py:\s*(\w+)\]', str(val))
			if m:
				cols[m.group(1)] = col_idx
	return cols


def parse_range_str(s):
	"""Parse 'U+XXXX-U+YYYY' into (start_int, end_int) or None."""
	m = re.match(r'U\+([0-9A-Fa-f]+)\s*-\s*U\+([0-9A-Fa-f]+)', str(s).strip())
	if m:
		return (int(m.group(1), 16), int(m.group(2), 16))
	return None


def build_row_index(adapter, range_col):
	"""Build dict: (start, end) -> row_number for all existing ranges.
	   Stops after 100 consecutive empty rows to avoid scanning padding."""
	idx = {}
	empty_run = 0
	for row_idx in range(2, adapter.max_row + 1):
		val = adapter.cell_value(row_idx, range_col)
		if val:
			empty_run = 0
			parsed = parse_range_str(val)
			if parsed:
				idx[parsed] = row_idx
		else:
			empty_run += 1
			if empty_run > 100:
				break
	return idx


def find_matching_row(chunk_start, chunk_end, block_start, block_end, row_index):
	"""Find existing row matching this chunk. Returns row_number or None."""
	# Exact match on chunk range
	if (chunk_start, chunk_end) in row_index:
		return row_index[(chunk_start, chunk_end)]
	# Subset match: any existing range within block that overlaps this chunk
	for (rs, re_), row in row_index.items():
		if block_start <= rs and re_ <= block_end:
			if rs <= chunk_end and re_ >= chunk_start:
				return row
	return None


def generate_block_chunks():
	"""Generate all (range_str, chunk_start, chunk_end, orig_str, block_start, block_end)
	   tuples for blocks below MAX_CODEPOINT."""
	blocks = parse_blocks(BLOCKS_DATA)
	chunks_out = []
	for start, end, name in blocks:
		if start >= MAX_CODEPOINT:
			continue
		chars = []
		for cp in range(start, end + 1):
			if not is_printable(cp):
				continue
			c = chr(cp)
			if unicodedata.category(c).startswith('Z'):
				continue
			chars.append(c)
		if not chars:
			continue
		for chunk in split_chunks(chars, MAX_ROW):
			chunk_start = ord(chunk[0])
			chunk_end = ord(chunk[-1])
			range_str = fmt_range(chunk_start, chunk_end)
			orig_str = ' '.join(chunk)
			chunks_out.append((range_str, chunk_start, chunk_end, orig_str, start, end))
	return chunks_out


def process_spreadsheet(path, chunks):
	"""Process a single spreadsheet file with the given block chunks."""
	print(f"\nOpening {path}", file=sys.stderr)
	adapter = open_spreadsheet(path)

	# Discover columns
	py_cols = find_py_columns(adapter)
	missing = [t for t in REQUIRED_TAGS if t not in py_cols]
	if missing:
		print(f"Error: missing required column tags in {os.path.basename(path)}: {missing}", file=sys.stderr)
		print(f"Found tags: {py_cols}", file=sys.stderr)
		return False

	col_range   = py_cols['range']
	col_orig    = py_cols['orig']
	col_junk    = py_cols['filter_junk']
	col_messy   = py_cols['filter_messy']
	col_visual  = py_cols['filter_visual']
	col_filt_out= py_cols['filtered_out']
	col_updated = py_cols['updated']
	print(f"Columns: range={col_range}, orig={col_orig}, filter_junk={col_junk}, filter_messy={col_messy}, filter_visual={col_visual}, filtered_out={col_filt_out}, updated={col_updated}", file=sys.stderr)

	# Build row index
	row_index = build_row_index(adapter, col_range)
	print(f"Found {len(row_index)} existing range rows", file=sys.stderr)

	# Find actual last data row for appending
	next_row = max(row_index.values()) + 1 if row_index else 2

	for i, (range_str, chunk_start, chunk_end, orig_str, block_start, block_end) in enumerate(chunks):
		# Find existing row or allocate new one
		row = find_matching_row(chunk_start, chunk_end, block_start, block_end, row_index)
		is_new = row is None
		if is_new:
			row = next_row
			next_row += 1

		# [py: range] — populate only if empty
		existing_range = adapter.cell_value(row, col_range)
		if not existing_range:
			adapter.set_cell_value(row, col_range, range_str)

		# [py: orig] — populate only if empty
		existing_orig = adapter.cell_value(row, col_orig)
		if not existing_orig:
			adapter.set_cell_value(row, col_orig, orig_str)
			# Update range based on actual filtered chars
			actual_chars = orig_str.split()
			if actual_chars:
				actual_start = ord(actual_chars[0])
				actual_end = ord(actual_chars[-1])
				adapter.set_cell_value(row, col_range, fmt_range(actual_start, actual_end))

		# Read current orig for filter inputs
		current_orig = adapter.cell_value(row, col_orig) or ''

		# [py: filter_junk] — always recompute
		if current_orig:
			junk_result = filter_junk_extract(current_orig)
			adapter.set_cell_value(row, col_junk, junk_result)
		else:
			junk_result = ''
			adapter.set_cell_value(row, col_junk, '')

		# [py: filter_messy] — always recompute
		if junk_result:
			messy_result = filter_messy_extract(junk_result)
			adapter.set_cell_value(row, col_messy, messy_result)
		else:
			messy_result = ''
			adapter.set_cell_value(row, col_messy, '')

		# [py: filter_visual] — always recompute
		if messy_result:
			visual_result = filter_visual_extract(messy_result)
			adapter.set_cell_value(row, col_visual, visual_result)
		else:
			adapter.set_cell_value(row, col_visual, '')

		# [py: filtered_out] — characters in orig but not in filter_visual
		orig_chars = set(current_orig.split()) if current_orig else set()
		visual_str = visual_result if messy_result else (adapter.cell_value(row, col_visual) or '')
		visual_chars = set(visual_str.split()) if visual_str else set()
		removed = [c for c in current_orig.split() if c not in visual_chars] if current_orig else []
		adapter.set_cell_value(row, col_filt_out, ' '.join(removed) if removed else '')

		# [py: updated] — timestamp this row
		adapter.set_cell_value(row, col_updated, datetime.now())
		adapter.set_cell_number_format(row, col_updated, 'YYYY-MM-DD HH:MM:SS')

		status = "NEW" if is_new else "updated"
		print(f"  [{i+1}/{len(chunks)}] Row {row} ({status}): {range_str}", file=sys.stderr)

	# Save
	adapter.save()
	print(f"Saved {path}", file=sys.stderr)
	return True


def main():
	# Generate block chunks once (same data for all spreadsheets)
	chunks = generate_block_chunks()
	print(f"Generated {len(chunks)} block chunks below U+{MAX_CODEPOINT:05X}", file=sys.stderr)

	failed = []
	for name in SPREADSHEET_NAMES:
		path = os.path.join(SPREADSHEET_DIR, name)
		if not os.path.exists(path):
			print(f"\nSkipping {name} (file not found)", file=sys.stderr)
			continue
		if not process_spreadsheet(path, chunks):
			failed.append(name)

	if failed:
		print(f"\nFailed: {', '.join(failed)}", file=sys.stderr)
		sys.exit(1)
	print(f"\nDone. Processed {len(SPREADSHEET_NAMES) - len(failed)} spreadsheet(s).", file=sys.stderr)


if __name__ == '__main__':
	main()

